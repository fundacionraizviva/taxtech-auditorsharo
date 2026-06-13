import streamlit as st
import pandas as pd
import numpy as np
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.series import SeriesLabel
import plotly.express as px

st.set_page_config(
    page_title="TaxTech Auditor RD",
    page_icon="🏛️",
    layout="wide",
    initial_sidebar_state="expanded"
)

# ─── ESTILOS CSS ──────────────────────────────────────────────────────────────
st.markdown("""
<style>
    [data-testid="stSidebar"] { background: #0f1923; }
    [data-testid="stSidebar"] * { color: #e8edf2 !important; }
    [data-testid="stSidebar"] .stSelectbox label,
    [data-testid="stSidebar"] .stTextInput label,
    [data-testid="stSidebar"] .stSlider label { color: #94a3b8 !important; font-size: 0.78rem !important; }
    [data-testid="stSidebar"] h1 { color: #38bdf8 !important; font-size: 1rem !important; letter-spacing: 0.05em; text-transform: uppercase; }
    .block-container { padding-top: 1.5rem; }
    div[data-testid="metric-container"] { background: #f8fafc; border: 1px solid #e2e8f0; border-radius: 8px; padding: 1rem; }
    div[data-testid="metric-container"] label { font-size: 0.75rem !important; color: #64748b !important; text-transform: uppercase; letter-spacing: 0.06em; }
    .stTabs [data-baseweb="tab"] { font-size: 0.82rem; padding: 6px 14px; }
    .stTabs [aria-selected="true"] { border-bottom: 2px solid #1e3a5f; font-weight: 700; }
    h1 { color: #1e3a5f !important; }
    h3 { color: #1e3a5f !important; font-size: 1rem !important; }
    
    .tabla-contable { width: 100%; border-collapse: collapse; font-family: 'Calibri', sans-serif; font-size: 0.9rem; margin-bottom: 1rem; }
    .tabla-contable th { border-bottom: 1px solid #000; padding: 8px; text-align: right; font-weight: bold; }
    .tabla-contable th:first-child { text-align: left; }
    .tabla-contable td { padding: 6px 8px; text-align: right; }
    .tabla-contable td:first-child { text-align: left; }
    .tabla-contable .seccion { font-weight: bold; text-align: left; padding-top: 15px; }
    .tabla-contable .total td { border-top: 1px solid #000; border-bottom: 3px double #000; font-weight: bold; }
    .tabla-contable .subtotal td { border-top: 1px solid #ccc; font-weight: bold; }
    .tabla-contable .titulo-anio { text-align: center; font-weight: bold; font-size: 0.95rem; background-color: #f1f5f9; padding: 8px !important; border-bottom: 2px solid #000;}
    
    .tabla-ir2 { width: 100%; border-collapse: collapse; font-family: 'Calibri', sans-serif; font-size: 0.85rem; margin-bottom: 2rem; border: 1px solid #cbd5e1; }
    .tabla-ir2 th { background-color: #1e3a5f; color: white; padding: 10px; text-align: center; font-weight: bold; border: 1px solid #cbd5e1; }
    .tabla-ir2 .header-seccion { background-color: #f1f5f9; font-weight: bold; color: #0f1923; text-align: left; padding: 6px; border: 1px solid #cbd5e1; font-size: 0.9rem; }
    .tabla-ir2 td { padding: 6px; border: 1px solid #cbd5e1; }
    .tabla-ir2 .col-num { width: 6%; text-align: center; font-weight: bold; background-color: #f8fafc; color: #475569; }
    .tabla-ir2 .col-desc { width: 64%; }
    .tabla-ir2 .col-monto { width: 30%; text-align: right; font-family: monospace; font-size:0.95rem;}
    .tabla-ir2 .fila-total td { font-weight: bold; background-color: #e2e8f0; }
    
    .alerta-box { background: #fff7ed; border-left: 4px solid #f59e0b; padding: 10px 14px; border-radius: 4px; margin: 8px 0; font-size: 0.85rem; }
    .info-box { background: #eff6ff; border-left: 4px solid #3b82f6; padding: 10px 14px; border-radius: 4px; margin: 8px 0; font-size: 0.85rem; }
    .ok-box { background: #f0fdf4; border-left: 4px solid #22c55e; padding: 10px 14px; border-radius: 4px; margin: 8px 0; font-size: 0.85rem; }
</style>
""", unsafe_allow_html=True)

# ──────────────────────────────────────────────────────────────────────────────
# PARÁMETROS FISCALES Y CLASIFICACIÓN
# ──────────────────────────────────────────────────────────────────────────────
NATURALEZAS = {'1': 'Debito', '2': 'Credito', '3': 'Credito', '4': 'Credito', '5': 'Debito', '6': 'Debito'}

PALABRAS_CRITICAS_ART287 = {
    'combustible': 'Art. 287 CTRD: Validar NCF válido y medios de pago para crédito ITBIS.',
    'representacion': 'Art. 287 CTRD: Razonabilidad, proporcionalidad y documentación fehaciente.',
    'retribucion': 'Art. 318 / Reg. 139-98: Retribuciones en especie. Validar ISR sustitutivo.',
    'gasto de personal': 'Art. 287 CTRD: Cruce obligatorio con IR-4 (TSS) para admitir deducción.',
    'honorario': 'Art. 309 CTRD: Retención 10% personas físicas / 2% entre jurídicas.',
}

TASA_ITBIS = 0.18
TASA_SFS_PAT = 0.0709; TASA_AFP_PAT = 0.0710
TASA_SRL = 0.0120; TASA_INFOTEP = 0.0100
TASA_SFS_EMP = 0.0304; TASA_AFP_EMP = 0.0287

# Techos TSS vigentes
TECHO_SFS = 118476.99
TECHO_AFP = 118476.99

def fmt_c(val):
    if pd.isna(val) or round(val, 2) == 0: return "-"
    return f"({abs(val):,.0f})" if val < 0 else f"{val:,.0f}"

def es_activo_no_corriente(cod, nombre):
    c = str(cod); n = str(nombre).lower()
    if c.startswith(('15', '16', '17', '18', '19')): return True
    keywords = ['fijo', 'propiedad', 'planta', 'equipo', 'depreciacion', 'edificio', 
                'terreno', 'vehiculo', 'software', 'intangible', 'mejora', 'inversion']
    if any(k in n for k in keywords) and 'gasto' not in n: return True
    return False

def es_pasivo_no_corriente(cod, nombre):
    c = str(cod); n = str(nombre).lower()
    if c.startswith(('22', '23', '24')): return True
    if any(k in n for k in ['largo plazo', 'prestamo bancario', 'bono', 'hipoteca']): return True
    return False

# ──────────────────────────────────────────────────────────────────────────────
# LECTOR DE ARCHIVOS ROBUSTO (Tolerante a "Falsos Excel", CSVs y TXTs)
# ──────────────────────────────────────────────────────────────────────────────
def leer_archivo_robusto(file):
    file_bytes = file.read()
    file.seek(0)
    try:
        return pd.read_excel(io.BytesIO(file_bytes), header=None)
    except: pass
    try: text = file_bytes.decode('utf-8')
    except: text = file_bytes.decode('latin1', errors='ignore')
    lines = text.splitlines()
    if not lines: return pd.DataFrame()
    sep = ','
    for s in [';', '\t', '|']:
        if lines[0].count(s) > lines[0].count(sep):
            sep = s
    return pd.read_csv(io.StringIO(text), sep=sep, header=None, dtype=str)

# ──────────────────────────────────────────────────────────────────────────────
# PROCESAMIENTO BALANZA
# ──────────────────────────────────────────────────────────────────────────────
def procesar_balanza(file) -> pd.DataFrame:
    try:
        df_raw = leer_archivo_robusto(file)
        if df_raw is None or df_raw.empty: return pd.DataFrame()
        header_idx = 0
        for idx, row in df_raw.iterrows():
            row_str = ' '.join([str(x).lower() for x in row.values])
            if ('código' in row_str or 'codigo' in row_str) and ('nombre' in row_str or 'cuenta' in row_str):
                header_idx = idx; break
        column_names = df_raw.iloc[header_idx].astype(str).str.lower().str.strip()
        df = df_raw.iloc[header_idx + 1:].reset_index(drop=True)
        idx_codigo, idx_cuenta = -1, -1
        indices_debe, indices_haber, indices_balance = [], [], []
        for i, col in enumerate(column_names):
            if any(x in col for x in ['código', 'codigo', 'cuenta no']) and idx_codigo == -1: idx_codigo = i
            elif any(x in col for x in ['nombre', 'descripción', 'cuenta']) and 'codigo' not in col and idx_cuenta == -1: idx_cuenta = i
            elif any(x in col for x in ['débito', 'debito', 'debe', 'cargos']): indices_debe.append(i)
            elif any(x in col for x in ['crédito', 'credito', 'haber', 'abonos']): indices_haber.append(i)
            elif any(x in col for x in ['saldo', 'balance', 'final', 'monto']): indices_balance.append(i)
        if idx_codigo == -1 or idx_cuenta == -1: return pd.DataFrame()
        col_dict = {'codigo': df.iloc[:, idx_codigo], 'cuenta': df.iloc[:, idx_cuenta]}
        if indices_debe: col_dict['debito'] = df.iloc[:, indices_debe[-1]]
        if indices_haber: col_dict['credito'] = df.iloc[:, indices_haber[-1]]
        if indices_balance: col_dict['saldo_final'] = df.iloc[:, indices_balance[-1]]
        df_clean = pd.DataFrame(col_dict)
        if 'saldo_final' not in df_clean.columns and 'debito' in df_clean.columns:
            df_clean['saldo_final'] = pd.to_numeric(df_clean['debito'], errors='coerce').fillna(0) - pd.to_numeric(df_clean['credito'], errors='coerce').fillna(0)
        df_clean['codigo'] = df_clean['codigo'].fillna('').astype(str).str.strip().apply(lambda x: x.split('.')[0] if '.' in x else x)
        df_clean['cuenta'] = df_clean['cuenta'].fillna('').astype(str).str.strip()
        df_clean = df_clean[(df_clean['codigo'] != '') & (~df_clean['codigo'].str.lower().str.contains('total|suma', na=False))]
        for col in ['debito', 'credito', 'saldo_final']:
            if col in df_clean.columns:
                df_clean[col] = pd.to_numeric(df_clean[col].astype(str).str.replace(',', '').replace(r'^-*$', '0', regex=True), errors='coerce').fillna(0.0)
        return df_clean.reset_index(drop=True)
    except:
        return pd.DataFrame()

# ──────────────────────────────────────────────────────────────────────────────
# PROCESAMIENTO 606/607
# ──────────────────────────────────────────────────────────────────────────────
def procesar_606_607(file, tipo):
    try:
        df_raw = leer_archivo_robusto(file)
        if df_raw is None or df_raw.empty: return 0.0, 0.0, pd.DataFrame()
        h_idx = -1
        for idx, row in df_raw.iterrows():
            row_str = "".join([str(x).lower() for x in row.values])
            if ('rnc' in row_str or 'cedula' in row_str) and ('ncf' in row_str or 'monto' in row_str):
                h_idx = idx; break
        if h_idx == -1: return 0.0, 0.0, pd.DataFrame()
        cols = [str(c).lower().replace('í','i').replace('ó','o').replace('\n',' ').strip() for c in df_raw.iloc[h_idx].values]
        df = pd.DataFrame(df_raw.iloc[h_idx + 1:].values, columns=cols)
        df = df.dropna(how='all')
        col_monto, col_itbis = None, None
        for c in df.columns:
            if 'total monto facturado' in c or 'monto total' in c:
                col_monto = c; break
        if not col_monto:
            for c in df.columns:
                if 'monto facturado' in c and 'bienes' not in c and 'servicios' not in c:
                    col_monto = c; break
        if tipo == "606":
            col_itbis = next((c for c in df.columns if 'itbis por adelantar' in c), None)
            if not col_itbis: col_itbis = next((c for c in df.columns if 'itbis facturado' in c), None)
        else:
            col_itbis = next((c for c in df.columns if 'itbis facturado' in c or 'itbis cobrado' in c), None)
        if col_monto:
            df[col_monto] = pd.to_numeric(df[col_monto].astype(str).str.replace(',', ''), errors='coerce').fillna(0)
        if col_itbis:
            df[col_itbis] = pd.to_numeric(df[col_itbis].astype(str).str.replace(',', ''), errors='coerce').fillna(0)
        monto_total = df[col_monto].sum() if col_monto else 0.0
        itbis_total = df[col_itbis].sum() if col_itbis else 0.0
        return monto_total, itbis_total, df
    except:
        return 0.0, 0.0, pd.DataFrame()

# ──────────────────────────────────────────────────────────────────────────────
# PROCESAMIENTO TSS
# ──────────────────────────────────────────────────────────────────────────────
def procesar_tss(file):
    try:
        df_raw = leer_archivo_robusto(file)
        if df_raw is None or df_raw.empty: return None, None
        h_idx = 0
        for idx, row in df_raw.iterrows():
            row_str = ' '.join([str(x).lower() for x in row.values])
            if 'cédula' in row_str or 'cedula' in row_str:
                h_idx = idx; break
        df = pd.DataFrame(df_raw.iloc[h_idx + 1:].values, columns=df_raw.iloc[h_idx].astype(str).str.lower().str.strip())
        col_salario = next((c for c in df.columns if 'salario ordinario' in str(c) or 'sueldo' in str(c)), None)
        col_nombre  = next((c for c in df.columns if 'nombre' in str(c) or 'empleado' in str(c)), None)
        col_cedula  = next((c for c in df.columns if 'cédula' in str(c) or 'cedula' in str(c)), None)
        col_otras   = next((c for c in df.columns if 'otras remuneraci' in str(c) or 'otras remun' in str(c)), None)
        if not col_salario: return None, None
        df[col_salario] = pd.to_numeric(df[col_salario].astype(str).str.replace(',', ''), errors='coerce').fillna(0)
        if col_otras:
            df[col_otras] = pd.to_numeric(df[col_otras].astype(str).str.replace(',', ''), errors='coerce').fillna(0)
        df_valid = df[df[col_salario] > 0].copy()
        df_valid['_salario'] = df_valid[col_salario]
        df_valid['_otras'] = df_valid[col_otras] if col_otras else 0
        df_valid['_base_sfs'] = df_valid['_salario'].clip(upper=TECHO_SFS)
        df_valid['_base_afp'] = df_valid['_salario'].clip(upper=TECHO_AFP)
        df_valid['SFS Patronal']    = (df_valid['_base_sfs'] * TASA_SFS_PAT).round(2)
        df_valid['AFP Patronal']    = (df_valid['_base_afp'] * TASA_AFP_PAT).round(2)
        df_valid['SRL Patronal']    = (df_valid['_salario']  * TASA_SRL).round(2)
        df_valid['INFOTEP']         = (df_valid['_salario']  * TASA_INFOTEP).round(2)
        df_valid['SFS Empleado']    = (df_valid['_base_sfs'] * TASA_SFS_EMP).round(2)
        df_valid['AFP Empleado']    = (df_valid['_base_afp'] * TASA_AFP_EMP).round(2)
        df_valid['Total Patronal']  = df_valid['SFS Patronal'] + df_valid['AFP Patronal'] + df_valid['SRL Patronal'] + df_valid['INFOTEP']
        df_valid['Total Empleado']  = df_valid['SFS Empleado'] + df_valid['AFP Empleado']
        df_valid['Total Nomina']    = df_valid['_salario'] + df_valid['_otras']
        df_valid['Neto a Pagar']    = df_valid['Total Nomina'] - df_valid['Total Empleado']
        cols_out = []
        if col_cedula:  cols_out.append(col_cedula)
        if col_nombre:  cols_out.append(col_nombre)
        cols_out += [col_salario, 'SFS Patronal', 'AFP Patronal', 'SRL Patronal', 'INFOTEP',
                     'Total Patronal', 'SFS Empleado', 'AFP Empleado', 'Total Empleado', 'Neto a Pagar']
        df_out = df_valid[[c for c in cols_out if c in df_valid.columns]].copy()
        salario_total = df_valid['_salario'].sum()
        res = {
            'total_empleados': len(df_valid),
            'nomina_mensual': salario_total,
            'sfs_pat': df_valid['SFS Patronal'].sum(),
            'afp_pat': df_valid['AFP Patronal'].sum(),
            'srl_pat': df_valid['SRL Patronal'].sum(),
            'infotep': df_valid['INFOTEP'].sum(),
            'sfs_emp': df_valid['SFS Empleado'].sum(),
            'afp_emp': df_valid['AFP Empleado'].sum(),
            'total_patronal': df_valid['Total Patronal'].sum(),
            'total_empleado': df_valid['Total Empleado'].sum(),
        }
        res['total_pagar'] = res['total_patronal'] + res['total_empleado']
        return df_out, res
    except:
        return None, None

def generar_plantilla_tss():
    df = pd.DataFrame(columns=[
        "Tipo de Documento", "Cédula / Pasaporte", "Nombres", "Apellidos", "Sexo",
        "Fecha de Nacimiento", "Salario Ordinario", "Otras Remuneraciones", "Aporte Voluntario"
    ])
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine='openpyxl') as writer:
        df.to_excel(writer, index=False)
    return buf.getvalue()

def generar_excel_descargable(df, sheet_name='Datos'):
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_name)
    return buf.getvalue()

# ──────────────────────────────────────────────────────────────────────────────
# LÓGICA DE AUDITORÍA
# ──────────────────────────────────────────────────────────────────────────────
def analizar_balanza(df: pd.DataFrame) -> pd.DataFrame:
    alertas_nat, alertas_fisc = [], []
    for row in df.itertuples(index=False):
        cod, nom, saldo = str(row.codigo), str(row.cuenta).lower(), row.saldo_final
        nat_esp = NATURALEZAS.get(cod[0] if cod else '', None)
        if any(x in nom for x in ['acum', 'depreciacion acum', 'deterioro', 'provision', 'amortizacion acum']):
            nat_esp = 'Credito' if nat_esp == 'Debito' else 'Debito'
        if nat_esp == 'Debito' and saldo < -1: alertas_nat.append("⚠️ Saldo crédito (nat. débito)")
        elif nat_esp == 'Credito' and saldo > 1: alertas_nat.append("⚠️ Saldo débito (nat. crédito)")
        else: alertas_nat.append("✅ Correcto")
        alertas_fisc.append(next((msg for p, msg in PALABRAS_CRITICAS_ART287.items() if p in nom), ""))
    df['validacion_naturaleza'] = alertas_nat
    df['alerta_fiscal'] = alertas_fisc
    return df

def procesar_comparativo(df_act: pd.DataFrame, df_ant: pd.DataFrame) -> pd.DataFrame:
    df_comp = pd.merge(df_act[['codigo', 'cuenta', 'saldo_final']], df_ant[['codigo', 'saldo_final']], on='codigo', how='outer', suffixes=('_Y2', '_Y1')).fillna(0.0)
    df_comp.loc[df_comp['cuenta'] == 0.0, 'cuenta'] = "Cuenta Histórica"
    df_comp['variacion_abs'] = df_comp['saldo_final_Y2'] - df_comp['saldo_final_Y1']
    df_comp['variacion_pct'] = np.where(df_comp['saldo_final_Y1'] != 0, (df_comp['variacion_abs'] / df_comp['saldo_final_Y1'].replace(0, np.nan)), 0)
    return df_comp

def calcular_casillas_ir2(df: pd.DataFrame) -> dict:
    def suma(prefijos):
        mask = df['codigo'].apply(lambda x: any(str(x).startswith(p) for p in prefijos))
        return abs(df.loc[mask, 'saldo_final'].sum())
    total_pasivos  = suma(['2'])
    inventario     = suma(['13', '130', '131', '132', '133'])
    af1 = suma(['152', '153']); af2 = suma(['154', '155']); af3 = suma(['156', '157', '158'])
    otros_act      = suma(['14', '16', '17', '18', '19'])
    saldo_act_fiscal = af1 + af2 + af3 + inventario + otros_act
    patrimonio_fisc  = saldo_act_fiscal - total_pasivos
    total_no_monet   = af1 + af2 + af3 + inventario
    return {'cas_34': min(max(patrimonio_fisc, 0), total_no_monet)}

# ──────────────────────────────────────────────────────────────────────────────
# EXPORTADORES EXCEL
# ──────────────────────────────────────────────────────────────────────────────
def exportar_reporte_corporativo(empresa, periodo, anio_act, df_comp):
    try:
        wb = openpyxl.Workbook()
        FNT_TITLE = Font(name="Calibri", size=14, bold=True, color="1F497D")
        FNT_HDR   = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        FNT_B     = Font(name="Calibri", size=11, bold=True)
        FNT_N     = Font(name="Calibri", size=11)
        FNT_POS   = Font(name="Calibri", size=11, color="166534")
        FNT_NEG   = Font(name="Calibri", size=11, color="991B1B")
        FILL_HDR  = PatternFill("solid", fgColor="1F497D")
        FILL_SUB  = PatternFill("solid", fgColor="F1F5F9")
        FILL_TOT  = PatternFill("solid", fgColor="E2E8F0")
        FILL_DASH = PatternFill("solid", fgColor="0F1923")
        THIN      = Side(border_style="thin", color="CBD5E1")
        B_BTM     = Border(bottom=THIN)
        B_DBL     = Border(top=THIN, bottom=Side(border_style="double", color="000000"))
        FMT_ACC   = '_(* #,##0.00_);_(* (#,##0.00);_(* "-"??_);_(@_)'
        FMT_PCT   = '0.00%'
        anio_prev = int(anio_act) - 1

        def _pct(y2, y1): return (y2 - y1) / abs(y1) if y1 != 0 else (1.0 if y2 != 0 else 0.0)

        def format_row(ws, row_n, style_type='normal'):
            for cell in ws[row_n]:
                cell.font = FNT_B if style_type in ['sub', 'sec', 'tot'] else FNT_N
                if style_type == 'sec': cell.fill = FILL_SUB; cell.border = B_BTM
                if style_type == 'tot': cell.fill = FILL_TOT; cell.border = B_DBL
                if isinstance(cell.value, (int, float)):
                    cell.number_format = FMT_PCT if cell.column == 6 else FMT_ACC

        def color_deviation(cell_abs, cell_pct, v2, v1, higher_is_good=True):
            diff = v2 - v1
            good = (diff >= 0) if higher_is_good else (diff <= 0)
            f = FNT_POS if good else FNT_NEG
            cell_abs.font = f; cell_pct.font = f

        def create_sheet_header(ws, title, cols=None):
            default_cols = [(45,"Cuenta"),(15,"Nota"),(20,f"Año {anio_act}"),(20,f"Año {anio_prev}"),(18,"Variación RD$"),(14,"Variación %")]
            cols = cols or default_cols
            for i, (w, _) in enumerate(cols): ws.column_dimensions[get_column_letter(i+1)].width = w
            ws["A1"] = empresa.upper(); ws["A2"] = title
            ws["A3"] = f"Comparativo años fiscales {anio_act} vs {anio_prev} | {periodo}"
            for row_n in range(1,4): ws[f"A{row_n}"].font = FNT_TITLE
            for i, (_, h) in enumerate(cols, 1):
                c = ws.cell(row=5, column=i, value=h)
                c.font = FNT_HDR; c.fill = FILL_HDR; c.alignment = Alignment(horizontal="center", wrap_text=True)
            ws.row_dimensions[5].height = 30
            return 6

        # 1. BALANCE GENERAL
        ws_bg = wb.active; ws_bg.title = "Balance General"
        r = create_sheet_header(ws_bg, "ESTADO DE SITUACIÓN FINANCIERA")

        def process_section(ws, r_idx, title, prefix, is_current, is_asset, higher_good=True):
            ws.cell(row=r_idx, column=1, value=title); format_row(ws, r_idx, 'sec'); r_idx += 1
            tot_y2, tot_y1 = 0, 0
            for _, row in df_comp[df_comp['codigo'].str.startswith(prefix, na=False)].iterrows():
                check_nc = (es_activo_no_corriente if is_asset else es_pasivo_no_corriente)(row['codigo'], row['cuenta'])
                if (is_current and not check_nc) or (not is_current and check_nc):
                    v2, v1 = abs(row['saldo_final_Y2']), abs(row['saldo_final_Y1'])
                    if is_asset and 'acum' in str(row['cuenta']).lower(): v2, v1 = -v2, -v1
                    if v2 == 0 and v1 == 0: continue
                    tot_y2 += v2; tot_y1 += v1
                    ws.cell(row=r_idx, column=1, value=row['cuenta'].title())
                    ws.cell(row=r_idx, column=3, value=v2); ws.cell(row=r_idx, column=4, value=v1)
                    c_abs = ws.cell(row=r_idx, column=5, value=v2 - v1)
                    c_pct = ws.cell(row=r_idx, column=6, value=_pct(v2,v1)); c_pct.number_format = FMT_PCT
                    color_deviation(c_abs, c_pct, v2, v1, higher_good); format_row(ws, r_idx, 'normal'); r_idx += 1
            lbl = f"Total {title.lower().replace(':','').strip()}"
            ws.cell(row=r_idx, column=1, value=lbl)
            ws.cell(row=r_idx, column=3, value=tot_y2); ws.cell(row=r_idx, column=4, value=tot_y1)
            c_abs = ws.cell(row=r_idx, column=5, value=tot_y2-tot_y1)
            c_pct = ws.cell(row=r_idx, column=6, value=_pct(tot_y2,tot_y1)); c_pct.number_format = FMT_PCT
            color_deviation(c_abs, c_pct, tot_y2, tot_y1, higher_good); format_row(ws, r_idx, 'sub'); r_idx += 1
            return r_idx, tot_y2, tot_y1

        r, ac_y2,  ac_y1  = process_section(ws_bg, r, "Activos corrientes:",     '1', True,  True,  True)
        r, anc_y2, anc_y1 = process_section(ws_bg, r, "Activos no corrientes:",  '1', False, True,  True)
        tot_act_y2 = ac_y2 + anc_y2; tot_act_y1 = ac_y1 + anc_y1
        ws_bg.cell(row=r, column=1, value="TOTAL ACTIVOS")
        ws_bg.cell(row=r, column=3, value=tot_act_y2); ws_bg.cell(row=r, column=4, value=tot_act_y1)
        c_abs = ws_bg.cell(row=r, column=5, value=tot_act_y2-tot_act_y1)
        c_pct = ws_bg.cell(row=r, column=6, value=_pct(tot_act_y2,tot_act_y1)); c_pct.number_format = FMT_PCT
        color_deviation(c_abs, c_pct, tot_act_y2, tot_act_y1); format_row(ws_bg, r, 'tot'); r += 2

        r, pc_y2, pc_y1 = process_section(ws_bg, r, "Pasivos corrientes:",    '2', True,  False, False)
        r, pnc_y2,pnc_y1= process_section(ws_bg, r, "Pasivos no corrientes:", '2', False, False, False)
        ws_bg.cell(row=r, column=1, value="Patrimonio:"); format_row(ws_bg, r, 'sec'); r += 1
        pat_y2, pat_y1 = 0, 0
        for _, row in df_comp[df_comp['codigo'].str.startswith('3', na=False)].iterrows():
            v2, v1 = abs(row['saldo_final_Y2']), abs(row['saldo_final_Y1'])
            if v2 == 0 and v1 == 0: continue
            pat_y2 += v2; pat_y1 += v1
            ws_bg.cell(row=r, column=1, value=row['cuenta'].title())
            ws_bg.cell(row=r, column=3, value=v2); ws_bg.cell(row=r, column=4, value=v1)
            c_abs = ws_bg.cell(row=r, column=5, value=v2-v1)
            c_pct = ws_bg.cell(row=r, column=6, value=_pct(v2,v1)); c_pct.number_format = FMT_PCT
            color_deviation(c_abs, c_pct, v2, v1); format_row(ws_bg, r, 'normal'); r += 1
        ws_bg.cell(row=r, column=1, value="Total Patrimonio")
        ws_bg.cell(row=r, column=3, value=pat_y2); ws_bg.cell(row=r, column=4, value=pat_y1)
        c_abs = ws_bg.cell(row=r, column=5, value=pat_y2-pat_y1)
        c_pct = ws_bg.cell(row=r, column=6, value=_pct(pat_y2,pat_y1)); c_pct.number_format = FMT_PCT
        color_deviation(c_abs, c_pct, pat_y2, pat_y1); format_row(ws_bg, r, 'sub'); r += 1
        tot_pap_y2 = pc_y2+pnc_y2+pat_y2; tot_pap_y1 = pc_y1+pnc_y1+pat_y1
        ws_bg.cell(row=r, column=1, value="TOTAL PASIVOS Y PATRIMONIO")
        ws_bg.cell(row=r, column=3, value=tot_pap_y2); ws_bg.cell(row=r, column=4, value=tot_pap_y1)
        c_abs = ws_bg.cell(row=r, column=5, value=tot_pap_y2-tot_pap_y1)
        c_pct = ws_bg.cell(row=r, column=6, value=_pct(tot_pap_y2,tot_pap_y1)); c_pct.number_format = FMT_PCT
        color_deviation(c_abs, c_pct, tot_pap_y2, tot_pap_y1); format_row(ws_bg, r, 'tot')

        # 2. ESTADO DE RESULTADOS
        ws_er = wb.create_sheet("Estado de Resultados")
        r = create_sheet_header(ws_er, "ESTADO DE RESULTADOS INTEGRALES")
        ws_er.cell(row=r, column=1, value="Ingresos operacionales:"); format_row(ws_er, r, 'sec'); r += 1
        ing_y2, ing_y1 = 0, 0
        for _, row in df_comp[df_comp['codigo'].str.startswith('4', na=False)].iterrows():
            v2, v1 = abs(row['saldo_final_Y2']), abs(row['saldo_final_Y1'])
            if v2 == 0 and v1 == 0: continue
            ing_y2 += v2; ing_y1 += v1
            ws_er.cell(row=r, column=1, value=row['cuenta'].title())
            ws_er.cell(row=r, column=3, value=v2); ws_er.cell(row=r, column=4, value=v1)
            c_abs = ws_er.cell(row=r, column=5, value=v2-v1)
            c_pct = ws_er.cell(row=r, column=6, value=_pct(v2,v1)); c_pct.number_format = FMT_PCT
            color_deviation(c_abs, c_pct, v2, v1); format_row(ws_er, r, 'normal'); r += 1
        ws_er.cell(row=r, column=1, value="Total Ingresos")
        ws_er.cell(row=r, column=3, value=ing_y2); ws_er.cell(row=r, column=4, value=ing_y1)
        c_abs = ws_er.cell(row=r, column=5, value=ing_y2-ing_y1)
        c_pct = ws_er.cell(row=r, column=6, value=_pct(ing_y2,ing_y1)); c_pct.number_format = FMT_PCT
        color_deviation(c_abs, c_pct, ing_y2, ing_y1); format_row(ws_er, r, 'sub'); r += 1

        ws_er.cell(row=r, column=1, value="Costos de ventas:"); format_row(ws_er, r, 'sec'); r += 1
        cos_y2, cos_y1 = 0, 0
        for _, row in df_comp[df_comp['codigo'].str.startswith('5', na=False)].iterrows():
            v2, v1 = abs(row['saldo_final_Y2']), abs(row['saldo_final_Y1'])
            if v2 == 0 and v1 == 0: continue
            cos_y2 += v2; cos_y1 += v1
            ws_er.cell(row=r, column=1, value=row['cuenta'].title())
            ws_er.cell(row=r, column=3, value=-v2); ws_er.cell(row=r, column=4, value=-v1)
            c_abs = ws_er.cell(row=r, column=5, value=-(v2-v1))
            c_pct = ws_er.cell(row=r, column=6, value=_pct(v2,v1)); c_pct.number_format = FMT_PCT
            color_deviation(c_abs, c_pct, v1, v2); format_row(ws_er, r, 'normal'); r += 1
        ws_er.cell(row=r, column=1, value="Total Costos de Ventas")
        ws_er.cell(row=r, column=3, value=-cos_y2); ws_er.cell(row=r, column=4, value=-cos_y1)
        c_abs = ws_er.cell(row=r, column=5, value=-(cos_y2-cos_y1))
        c_pct = ws_er.cell(row=r, column=6, value=_pct(cos_y2,cos_y1)); c_pct.number_format = FMT_PCT
        color_deviation(c_abs, c_pct, cos_y1, cos_y2); format_row(ws_er, r, 'sub'); r += 1
        ub_y2 = ing_y2-cos_y2; ub_y1 = ing_y1-cos_y1
        ws_er.cell(row=r, column=1, value="UTILIDAD BRUTA")
        ws_er.cell(row=r, column=3, value=ub_y2); ws_er.cell(row=r, column=4, value=ub_y1)
        c_abs = ws_er.cell(row=r, column=5, value=ub_y2-ub_y1)
        c_pct = ws_er.cell(row=r, column=6, value=_pct(ub_y2,ub_y1)); c_pct.number_format = FMT_PCT
        color_deviation(c_abs, c_pct, ub_y2, ub_y1); format_row(ws_er, r, 'tot'); r += 1

        ws_er.cell(row=r, column=1, value="Gastos operacionales:"); format_row(ws_er, r, 'sec'); r += 1
        gas_y2, gas_y1 = 0, 0
        for _, row in df_comp[df_comp['codigo'].str.startswith('6', na=False)].iterrows():
            v2, v1 = abs(row['saldo_final_Y2']), abs(row['saldo_final_Y1'])
            if v2 == 0 and v1 == 0: continue
            gas_y2 += v2; gas_y1 += v1
            ws_er.cell(row=r, column=1, value=row['cuenta'].title())
            ws_er.cell(row=r, column=3, value=-v2); ws_er.cell(row=r, column=4, value=-v1)
            c_abs = ws_er.cell(row=r, column=5, value=-(v2-v1))
            c_pct = ws_er.cell(row=r, column=6, value=_pct(v2,v1)); c_pct.number_format = FMT_PCT
            color_deviation(c_abs, c_pct, v1, v2); format_row(ws_er, r, 'normal'); r += 1
        ws_er.cell(row=r, column=1, value="Total Gastos Operacionales")
        ws_er.cell(row=r, column=3, value=-gas_y2); ws_er.cell(row=r, column=4, value=-gas_y1)
        c_abs = ws_er.cell(row=r, column=5, value=-(gas_y2-gas_y1))
        c_pct = ws_er.cell(row=r, column=6, value=_pct(gas_y2,gas_y1)); c_pct.number_format = FMT_PCT
        color_deviation(c_abs, c_pct, gas_y1, gas_y2); format_row(ws_er, r, 'sub'); r += 1
        un_y2 = ing_y2-cos_y2-gas_y2; un_y1 = ing_y1-cos_y1-gas_y1
        ws_er.cell(row=r, column=1, value="UTILIDAD (PÉRDIDA) NETA DEL PERÍODO")
        ws_er.cell(row=r, column=3, value=un_y2); ws_er.cell(row=r, column=4, value=un_y1)
        c_abs = ws_er.cell(row=r, column=5, value=un_y2-un_y1)
        c_pct = ws_er.cell(row=r, column=6, value=_pct(un_y2,un_y1)); c_pct.number_format = FMT_PCT
        color_deviation(c_abs, c_pct, un_y2, un_y1); format_row(ws_er, r, 'tot')

        # 3. PATRIMONIO
        ws_pat = wb.create_sheet("Patrimonio")
        r = create_sheet_header(ws_pat, "ESTADO DE CAMBIOS EN EL PATRIMONIO")
        ws_pat.cell(row=r, column=1, value="Saldos y Movimientos del Período:"); format_row(ws_pat, r, 'sec'); r += 1
        for _, row in df_comp[df_comp['codigo'].str.startswith('3', na=False)].iterrows():
            v2, v1 = abs(row['saldo_final_Y2']), abs(row['saldo_final_Y1'])
            if v2 == 0 and v1 == 0: continue
            ws_pat.cell(row=r, column=1, value=row['cuenta'].title())
            ws_pat.cell(row=r, column=3, value=v2); ws_pat.cell(row=r, column=4, value=v1)
            c_abs = ws_pat.cell(row=r, column=5, value=v2-v1)
            c_pct = ws_pat.cell(row=r, column=6, value=_pct(v2,v1)); c_pct.number_format = FMT_PCT
            color_deviation(c_abs, c_pct, v2, v1); format_row(ws_pat, r, 'normal'); r += 1
        ws_pat.cell(row=r, column=1, value="TOTAL PATRIMONIO")
        ws_pat.cell(row=r, column=3, value=pat_y2); ws_pat.cell(row=r, column=4, value=pat_y1)
        c_abs = ws_pat.cell(row=r, column=5, value=pat_y2-pat_y1)
        c_pct = ws_pat.cell(row=r, column=6, value=_pct(pat_y2,pat_y1)); c_pct.number_format = FMT_PCT
        color_deviation(c_abs, c_pct, pat_y2, pat_y1); format_row(ws_pat, r, 'tot')

        # 4. FLUJO DE EFECTIVO
        ws_fe = wb.create_sheet("Flujo de Efectivo")
        r = create_sheet_header(ws_fe, "ESTADO DE FLUJO DE EFECTIVO (Método Indirecto)")

        def fe_row(ws, r_idx, lbl, v2, v1, higher_good=True, style='normal'):
            ws.cell(row=r_idx, column=1, value=lbl)
            ws.cell(row=r_idx, column=3, value=v2); ws.cell(row=r_idx, column=4, value=v1)
            c_abs = ws.cell(row=r_idx, column=5, value=v2-v1)
            c_pct = ws.cell(row=r_idx, column=6, value=_pct(v2,v1)); c_pct.number_format = FMT_PCT
            color_deviation(c_abs, c_pct, v2, v1, higher_good); format_row(ws, r_idx, style)
            return r_idx + 1

        ws_fe.cell(row=r, column=1, value="I. ACTIVIDADES OPERACIONALES"); format_row(ws_fe, r, 'sec'); r += 1
        r = fe_row(ws_fe, r, "Utilidad neta del período", un_y2, un_y1)
        ws_fe.cell(row=r, column=1, value="Ajustes por partidas no monetarias:"); format_row(ws_fe, r, 'sec'); r += 1
        dep_y2 = abs(df_comp[df_comp['cuenta'].str.lower().str.contains('deprecia', na=False)]['saldo_final_Y2'].sum())
        dep_y1 = abs(df_comp[df_comp['cuenta'].str.lower().str.contains('deprecia', na=False)]['saldo_final_Y1'].sum())
        r = fe_row(ws_fe, r, "(+) Depreciación y amortización", dep_y2, dep_y1)
        ws_fe.cell(row=r, column=1, value="Cambios en capital de trabajo:"); format_row(ws_fe, r, 'sec'); r += 1
        for _, row in df_comp[df_comp['codigo'].str.startswith('1', na=False)].iterrows():
            if not es_activo_no_corriente(row['codigo'], row['cuenta']):
                v2, v1 = abs(row['saldo_final_Y2']), abs(row['saldo_final_Y1'])
                if v2 == 0 and v1 == 0: continue
                r = fe_row(ws_fe, r, f"  (Aumento)/Disminución: {row['cuenta'].title()}", -(v2-v1), 0, False)
        for _, row in df_comp[df_comp['codigo'].str.startswith('2', na=False)].iterrows():
            if not es_pasivo_no_corriente(row['codigo'], row['cuenta']):
                v2, v1 = abs(row['saldo_final_Y2']), abs(row['saldo_final_Y1'])
                if v2 == 0 and v1 == 0: continue
                r = fe_row(ws_fe, r, f"  Aumento/(Disminución): {row['cuenta'].title()}", v2-v1, 0)
        op_cft_y2 = un_y2+dep_y2; op_cft_y1 = un_y1+dep_y1
        r = fe_row(ws_fe, r, "Efectivo neto de actividades operacionales", op_cft_y2, op_cft_y1, style='tot')
        r += 1
        ws_fe.cell(row=r, column=1, value="II. ACTIVIDADES DE INVERSIÓN"); format_row(ws_fe, r, 'sec'); r += 1
        inv_y2 = 0
        for _, row in df_comp[df_comp['codigo'].str.startswith('1', na=False)].iterrows():
            if es_activo_no_corriente(row['codigo'], row['cuenta']) and 'acum' not in str(row['cuenta']).lower():
                v2, v1 = abs(row['saldo_final_Y2']), abs(row['saldo_final_Y1'])
                if v2 == 0 and v1 == 0: continue
                fe2 = -(v2-v1); inv_y2 += fe2
                r = fe_row(ws_fe, r, f"  Adquisición/(Venta): {row['cuenta'].title()}", fe2, 0, False)
        r = fe_row(ws_fe, r, "Efectivo neto de actividades de inversión", inv_y2, 0, style='tot')
        r += 1
        ws_fe.cell(row=r, column=1, value="III. ACTIVIDADES DE FINANCIAMIENTO"); format_row(ws_fe, r, 'sec'); r += 1
        fin_y2 = 0
        for _, row in df_comp[df_comp['codigo'].str.startswith('2', na=False)].iterrows():
            if es_pasivo_no_corriente(row['codigo'], row['cuenta']):
                v2, v1 = abs(row['saldo_final_Y2']), abs(row['saldo_final_Y1'])
                if v2 == 0 and v1 == 0: continue
                fe2 = v2-v1; fin_y2 += fe2
                r = fe_row(ws_fe, r, f"  Variación: {row['cuenta'].title()}", fe2, 0)
        r = fe_row(ws_fe, r, "Efectivo neto de actividades de financiamiento", fin_y2, 0, style='tot')
        r += 1
        r = fe_row(ws_fe, r, "VARIACIÓN NETA EN EFECTIVO Y EQUIVALENTES", op_cft_y2+inv_y2+fin_y2, 0, style='tot')

        # 5. DASHBOARD
        ws_dash = wb.create_sheet("Dashboard", 0)
        ws_dash.sheet_properties.tabColor = "1F497D"
        ws_dash.column_dimensions['A'].width = 28
        for col in ['B','C','D','E','F','G','H','I','J','K','L','M','N','O','P']: ws_dash.column_dimensions[col].width = 10
        for row_n in range(1, 6):
            for col_n in range(1, 17): ws_dash.cell(row=row_n, column=col_n).fill = FILL_DASH
        ws_dash["B2"] = empresa.upper(); ws_dash["B2"].font = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
        ws_dash["B3"] = f"Dashboard Financiero Comparativo — {anio_act} vs {anio_prev}"
        ws_dash["B3"].font = Font(name="Calibri", size=11, color="94A3B8")
        ws_dash["B4"] = periodo; ws_dash["B4"].font = Font(name="Calibri", size=10, color="64748B")
        kpi_labels   = [f"Año {anio_act}", f"Año {anio_prev}", "Variación RD$", "Variación %"]
        kpi_concepts = ["Ingresos", "Utilidad Bruta", "Utilidad Neta", "Total Activos", "Total Pasivos", "Patrimonio"]
        kpi_data = [(ing_y2,ing_y1),(ub_y2,ub_y1),(un_y2,un_y1),(tot_act_y2,tot_act_y1),(pc_y2+pnc_y2,pc_y1+pnc_y1),(pat_y2,pat_y1)]
        ws_dash.cell(row=7, column=1, value="Indicador").font = FNT_HDR
        ws_dash.cell(row=7, column=1).fill = FILL_HDR
        for i, lbl in enumerate(kpi_labels, 2):
            c = ws_dash.cell(row=7, column=i, value=lbl)
            c.font = FNT_HDR; c.fill = FILL_HDR; c.alignment = Alignment(horizontal="center")
        for idx, (concept, (v2, v1)) in enumerate(zip(kpi_concepts, kpi_data), 8):
            ws_dash.cell(row=idx, column=1, value=concept).font = FNT_B
            ws_dash.cell(row=idx, column=2, value=v2).number_format = FMT_ACC
            ws_dash.cell(row=idx, column=3, value=v1).number_format = FMT_ACC
            c_dif = ws_dash.cell(row=idx, column=4, value=v2-v1)
            c_dif.number_format = FMT_ACC; c_dif.font = FNT_POS if v2 >= v1 else FNT_NEG
            c_pct_v = ws_dash.cell(row=idx, column=5, value=_pct(v2,v1))
            c_pct_v.number_format = FMT_PCT; c_pct_v.font = FNT_POS if v2 >= v1 else FNT_NEG
            if idx % 2 == 0:
                for col_n in range(1, 6): ws_dash.cell(row=idx, column=col_n).fill = FILL_SUB
        chart1 = BarChart(); chart1.type = "col"; chart1.grouping = "clustered"
        chart1.title = "P&L Comparativo"; chart1.style = 10
        chart1.y_axis.title = "RD$"; chart1.width = 16; chart1.height = 12
        for j, (lbl, v2, v1) in enumerate(zip(["Ingresos","Costo Ventas","Utilidad Bruta","Utilidad Neta"], [ing_y2,cos_y2,ub_y2,un_y2], [ing_y1,cos_y1,ub_y1,un_y1]), 8):
            ws_dash.cell(row=15, column=j, value=lbl)
            ws_dash.cell(row=16, column=j, value=v2); ws_dash.cell(row=17, column=j, value=v1)
        data_ref1 = Reference(ws_dash, min_col=8, max_col=11, min_row=16, max_row=17)
        cats_ref1 = Reference(ws_dash, min_col=8, max_col=11, min_row=15)
        chart1.add_data(data_ref1, from_rows=True); chart1.set_categories(cats_ref1)
        chart1.series[0].title = SeriesLabel(v=str(anio_act))
        chart1.series[1].title = SeriesLabel(v=str(anio_prev))
        ws_dash.add_chart(chart1, "A20")

        buf = io.BytesIO(); wb.save(buf)
        return buf.getvalue()
    except:
        return None

def exportar_balance_excel(df_comp, empresa, anio):
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Balance General"
    FNT_H = Font(bold=True, color="FFFFFF"); FILL_H = PatternFill("solid", fgColor="1F497D")
    FNT_B = Font(bold=True); FNT_N = Font()
    FILL_S = PatternFill("solid", fgColor="F1F5F9")
    FMT   = '#,##0.00'
    anio_p = int(anio) - 1
    ws.column_dimensions['A'].width = 45; ws.column_dimensions['B'].width = 18; ws.column_dimensions['C'].width = 18
    ws["A1"] = empresa.upper(); ws["A1"].font = Font(bold=True, size=14, color="1F497D")
    ws["A2"] = "ESTADO DE SITUACIÓN FINANCIERA"; ws["A2"].font = Font(bold=True, size=12)
    ws["A3"] = f"Al 31 de diciembre de {anio}"; ws["A3"].font = Font(italic=True)
    for col, hdr in [(1,"Cuenta"),(2,str(anio)),(3,str(anio_p))]:
        c = ws.cell(row=5, column=col, value=hdr); c.font = FNT_H; c.fill = FILL_H
    r = 6
    def sec(title):
        nonlocal r
        ws.cell(row=r, column=1, value=title).font = FNT_B
        ws.cell(row=r, column=1).fill = FILL_S; r += 1
    def fila(nombre, v2, v1):
        nonlocal r
        ws.cell(row=r, column=1, value=nombre).font = FNT_N
        ws.cell(row=r, column=2, value=v2).number_format = FMT
        ws.cell(row=r, column=3, value=v1).number_format = FMT; r += 1
        return v2, v1
    def subtot(nombre, t2, t1):
        nonlocal r
        for col, val in [(1,nombre),(2,t2),(3,t1)]:
            c = ws.cell(row=r, column=col, value=val); c.font = FNT_B
            if col > 1: c.number_format = FMT
        r += 1
    sec("ACTIVOS CORRIENTES")
    ac2, ac1 = 0, 0
    for _, row in df_comp[df_comp['codigo'].str.startswith('1', na=False)].iterrows():
        if not es_activo_no_corriente(row['codigo'], row['cuenta']):
            v2, v1 = abs(row['saldo_final_Y2']), abs(row['saldo_final_Y1'])
            if v2 or v1: fila(row['cuenta'].title(), v2, v1); ac2 += v2; ac1 += v1
    subtot("Total Activos Corrientes", ac2, ac1)
    sec("ACTIVOS NO CORRIENTES")
    anc2, anc1 = 0, 0
    for _, row in df_comp[df_comp['codigo'].str.startswith('1', na=False)].iterrows():
        if es_activo_no_corriente(row['codigo'], row['cuenta']):
            v2, v1 = abs(row['saldo_final_Y2']), abs(row['saldo_final_Y1'])
            if 'acum' in str(row['cuenta']).lower(): v2, v1 = -v2, -v1
            if v2 or v1: fila(row['cuenta'].title(), v2, v1); anc2 += v2; anc1 += v1
    subtot("Total Activos No Corrientes", anc2, anc1)
    for col, val in [(1,"TOTAL ACTIVOS"),(2,ac2+anc2),(3,ac1+anc1)]:
        c = ws.cell(row=r, column=col, value=val); c.font = Font(bold=True, size=12)
        if col > 1: c.number_format = FMT
    r += 2
    sec("PASIVOS CORRIENTES")
    pc2, pc1 = 0, 0
    for _, row in df_comp[df_comp['codigo'].str.startswith('2', na=False)].iterrows():
        if not es_pasivo_no_corriente(row['codigo'], row['cuenta']):
            v2, v1 = abs(row['saldo_final_Y2']), abs(row['saldo_final_Y1'])
            if v2 or v1: fila(row['cuenta'].title(), v2, v1); pc2 += v2; pc1 += v1
    subtot("Total Pasivos Corrientes", pc2, pc1)
    sec("PASIVOS NO CORRIENTES")
    pnc2, pnc1 = 0, 0
    for _, row in df_comp[df_comp['codigo'].str.startswith('2', na=False)].iterrows():
        if es_pasivo_no_corriente(row['codigo'], row['cuenta']):
            v2, v1 = abs(row['saldo_final_Y2']), abs(row['saldo_final_Y1'])
            if v2 or v1: fila(row['cuenta'].title(), v2, v1); pnc2 += v2; pnc1 += v1
    subtot("Total Pasivos No Corrientes", pnc2, pnc1)
    sec("PATRIMONIO")
    pat2, pat1 = 0, 0
    for _, row in df_comp[df_comp['codigo'].str.startswith('3', na=False)].iterrows():
        v2, v1 = abs(row['saldo_final_Y2']), abs(row['saldo_final_Y1'])
        if v2 or v1: fila(row['cuenta'].title(), v2, v1); pat2 += v2; pat1 += v1
    subtot("Total Patrimonio", pat2, pat1)
    for col, val in [(1,"TOTAL PASIVOS Y PATRIMONIO"),(2,pc2+pnc2+pat2),(3,pc1+pnc1+pat1)]:
        c = ws.cell(row=r, column=col, value=val); c.font = Font(bold=True, size=12)
        if col > 1: c.number_format = FMT
    buf = io.BytesIO(); wb.save(buf); return buf.getvalue()

def exportar_er_excel(df_comp, empresa, anio):
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Estado de Resultados"
    FNT_H = Font(bold=True, color="FFFFFF"); FILL_H = PatternFill("solid", fgColor="1F497D")
    FNT_B = Font(bold=True); FMT = '#,##0.00'; anio_p = int(anio) - 1
    FILL_S = PatternFill("solid", fgColor="F1F5F9")
    ws.column_dimensions['A'].width = 45; ws.column_dimensions['B'].width = 18; ws.column_dimensions['C'].width = 18
    ws["A1"] = empresa.upper(); ws["A1"].font = Font(bold=True, size=14, color="1F497D")
    ws["A2"] = "ESTADO DE RESULTADOS INTEGRALES"; ws["A2"].font = Font(bold=True, size=12)
    ws["A3"] = f"Por el período terminado el 31 de diciembre de {anio}"
    for col, hdr in [(1,"Concepto"),(2,str(anio)),(3,str(anio_p))]:
        c = ws.cell(row=5, column=col, value=hdr); c.font = FNT_H; c.fill = FILL_H
    r = 6
    def sec(t): 
        ws.cell(row=r, column=1, value=t).font = FNT_B
        ws.cell(row=r, column=1).fill = FILL_S
    def fila(n, v2, v1):
        nonlocal r
        ws.cell(row=r, column=1, value=n)
        ws.cell(row=r, column=2, value=v2).number_format = FMT
        ws.cell(row=r, column=3, value=v1).number_format = FMT; r += 1
    def tot(n, v2, v1):
        nonlocal r
        for col, val in [(1,n),(2,v2),(3,v1)]:
            c = ws.cell(row=r, column=col, value=val); c.font = FNT_B
            if col > 1: c.number_format = FMT
        r += 1
    sec("Ingresos:"); r += 1
    ing2, ing1 = 0, 0
    for _, row in df_comp[df_comp['codigo'].str.startswith('4', na=False)].iterrows():
        v2, v1 = abs(row['saldo_final_Y2']), abs(row['saldo_final_Y1'])
        if v2 or v1: fila(row['cuenta'].title(), v2, v1); ing2 += v2; ing1 += v1
    tot("Total Ingresos", ing2, ing1)
    sec("Costos:"); r += 1
    cos2, cos1 = 0, 0
    for _, row in df_comp[df_comp['codigo'].str.startswith('5', na=False)].iterrows():
        v2, v1 = abs(row['saldo_final_Y2']), abs(row['saldo_final_Y1'])
        if v2 or v1: fila(row['cuenta'].title(), -v2, -v1); cos2 += v2; cos1 += v1
    tot("Total Costos", -cos2, -cos1)
    tot("UTILIDAD BRUTA", ing2-cos2, ing1-cos1)
    sec("Gastos:"); r += 1
    gas2, gas1 = 0, 0
    for _, row in df_comp[df_comp['codigo'].str.startswith('6', na=False)].iterrows():
        v2, v1 = abs(row['saldo_final_Y2']), abs(row['saldo_final_Y1'])
        if v2 or v1: fila(row['cuenta'].title(), -v2, -v1); gas2 += v2; gas1 += v1
    tot("Total Gastos", -gas2, -gas1)
    tot("UTILIDAD (PÉRDIDA) NETA", ing2-cos2-gas2, ing1-cos1-gas1)
    buf = io.BytesIO(); wb.save(buf); return buf.getvalue()

def exportar_it1_excel(empresa, periodo, monto_607, itbis_607, monto_606, itbis_606, df_607=None, df_606=None):
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Preliquidación IT-1"
    FNT_H = Font(bold=True, color="FFFFFF"); FILL_H = PatternFill("solid", fgColor="1F497D")
    FNT_B = Font(bold=True); FMT = '#,##0.00'
    ws.column_dimensions['A'].width = 40; ws.column_dimensions['B'].width = 20
    ws["A1"] = empresa.upper(); ws["A1"].font = Font(bold=True, size=14, color="1F497D")
    ws["A2"] = "PRELIQUIDACIÓN IT-1 — ITBIS"; ws["A2"].font = Font(bold=True, size=12)
    ws["A3"] = f"Período: {periodo}"
    r = 5
    for col, h in [(1,"Concepto"),(2,"Monto RD$")]:
        c = ws.cell(row=r, column=col, value=h); c.font = FNT_H; c.fill = FILL_H
    r += 1
    data = [
        ("ITBIS COBRADO (Formato 607 — Ventas)", itbis_607),
        ("Monto Total Facturado (607)", monto_607),
        ("ITBIS ADELANTADO (Formato 606 — Compras)", itbis_606),
        ("Monto Total Compras (606)", monto_606),
        ("ITBIS A PAGAR / SALDO A FAVOR", itbis_607 - itbis_606),
    ]
    for label, val in data:
        ws.cell(row=r, column=1, value=label).font = FNT_B if "ITBIS" in label else Font()
        ws.cell(row=r, column=2, value=val).number_format = FMT; r += 1
    if df_607 is not None and not df_607.empty:
        ws2 = wb.create_sheet("Detalle 607")
        for c_idx, col in enumerate(df_607.columns, 1):
            ws2.cell(row=1, column=c_idx, value=col).font = FNT_H
            ws2.cell(row=1, column=c_idx).fill = FILL_H
        for r_idx, row_data in df_607.iterrows():
            for c_idx, val in enumerate(row_data, 1):
                ws2.cell(row=r_idx+2, column=c_idx, value=val)
    if df_606 is not None and not df_606.empty:
        ws3 = wb.create_sheet("Detalle 606")
        for c_idx, col in enumerate(df_606.columns, 1):
            ws3.cell(row=1, column=c_idx, value=col).font = FNT_H
            ws3.cell(row=1, column=c_idx).fill = FILL_H
        for r_idx, row_data in df_606.iterrows():
            for c_idx, val in enumerate(row_data, 1):
                ws3.cell(row=r_idx+2, column=c_idx, value=val)
    buf = io.BytesIO(); wb.save(buf); return buf.getvalue()

def exportar_tss_excel(df_tss, tss_res, empresa, periodo):
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Auditoría TSS"
    FNT_H = Font(bold=True, color="FFFFFF"); FILL_H = PatternFill("solid", fgColor="1F497D")
    FNT_B = Font(bold=True); FMT = '#,##0.00'
    ws["A1"] = empresa.upper(); ws["A1"].font = Font(bold=True, size=14, color="1F497D")
    ws["A2"] = "AUDITORÍA DE NÓMINA Y TSS"; ws["A2"].font = Font(bold=True, size=12)
    ws["A3"] = f"Período: {periodo}"
    r = 5
    if df_tss is not None and not df_tss.empty:
        for c_idx, col in enumerate(df_tss.columns, 1):
            c = ws.cell(row=r, column=c_idx, value=col); c.font = FNT_H; c.fill = FILL_H
            ws.column_dimensions[get_column_letter(c_idx)].width = 18
        r += 1
        for _, row_data in df_tss.iterrows():
            for c_idx, val in enumerate(row_data, 1):
                c = ws.cell(row=r, column=c_idx, value=val)
                if isinstance(val, (int, float)): c.number_format = FMT
            r += 1
    r += 2
    ws.cell(row=r, column=1, value="RESUMEN TSS").font = FNT_B
    r += 1
    if tss_res:
        for label, val in [
            ("Total Empleados", tss_res['total_empleados']),
            ("Nómina Total Mensual", tss_res['nomina_mensual']),
            ("Total Aportes Patronales", tss_res['total_patronal']),
            ("Total Retenciones Empleados", tss_res['total_empleado']),
            ("TOTAL A REMITIR TSS", tss_res['total_pagar']),
        ]:
            ws.cell(row=r, column=1, value=label).font = FNT_B
            c = ws.cell(row=r, column=2, value=val)
            if isinstance(val, float): c.number_format = FMT
            r += 1
    buf = io.BytesIO(); wb.save(buf); return buf.getvalue()

# ──────────────────────────────────────────────────────────────────────────────
# GENERADORES HTML
# ──────────────────────────────────────────────────────────────────────────────
def html_cambios_patrimonio(df_comp, anio):
    html = f"<table class='tabla-contable'><tr><th>Cuentas de Patrimonio</th><th>Saldo Inicial {int(anio)-1}</th><th>Variación</th><th>Saldo Final {anio}</th></tr>"
    pat_y2, pat_y1 = 0, 0
    for _, r in df_comp[df_comp['codigo'].str.startswith('3', na=False)].iterrows():
        v2, v1 = abs(r['saldo_final_Y2']), abs(r['saldo_final_Y1'])
        if v2 == 0 and v1 == 0: continue
        pat_y2 += v2; pat_y1 += v1
        html += f"<tr><td>{r['cuenta'].title()}</td><td>{fmt_c(v1)}</td><td>{fmt_c(v2-v1)}</td><td>{fmt_c(v2)}</td></tr>"
    html += f"<tr class='total'><td>TOTAL PATRIMONIO</td><td>{fmt_c(pat_y1)}</td><td>{fmt_c(pat_y2-pat_y1)}</td><td>{fmt_c(pat_y2)}</td></tr></table>"
    return html

def html_estado_resultados(df_comp, anio):
    html = f"<table class='tabla-contable'><tr><th>Conceptos</th><th>Nota</th><th>{anio}</th><th>{int(anio)-1}</th></tr>"
    html += "<tr><td class='seccion' colspan='4'>Ingresos operacionales:</td></tr>"
    ing_y2, ing_y1 = 0, 0
    for _, r in df_comp[df_comp['codigo'].str.startswith('4', na=False)].iterrows():
        v2, v1 = abs(r['saldo_final_Y2']), abs(r['saldo_final_Y1'])
        if v2 == 0 and v1 == 0: continue
        ing_y2 += v2; ing_y1 += v1
        html += f"<tr><td>{r['cuenta'].title()}</td><td></td><td>{fmt_c(v2)}</td><td>{fmt_c(v1)}</td></tr>"
    html += f"<tr class='subtotal'><td>Total Ingresos</td><td></td><td>{fmt_c(ing_y2)}</td><td>{fmt_c(ing_y1)}</td></tr>"
    html += "<tr><td class='seccion' colspan='4'>Costos y gastos operacionales:</td></tr>"
    gas_y2, gas_y1 = 0, 0
    for _, r in df_comp[df_comp['codigo'].str.startswith(('5','6'), na=False)].iterrows():
        v2, v1 = -abs(r['saldo_final_Y2']), -abs(r['saldo_final_Y1'])
        if v2 == 0 and v1 == 0: continue
        gas_y2 += v2; gas_y1 += v1
        html += f"<tr><td>{r['cuenta'].title()}</td><td></td><td>{fmt_c(v2)}</td><td>{fmt_c(v1)}</td></tr>"
    html += f"<tr class='subtotal'><td>Total Costos y Gastos</td><td></td><td>{fmt_c(gas_y2)}</td><td>{fmt_c(gas_y1)}</td></tr>"
    html += f"<tr class='total'><td>Utilidad (Pérdida) del Período</td><td></td><td>{fmt_c(ing_y2+gas_y2)}</td><td>{fmt_c(ing_y1+gas_y1)}</td></tr></table>"
    return html

def html_balance_general(df_comp, anio, tipo='activo'):
    html = f"<table class='tabla-contable'><tr><th>{tipo.capitalize()}s</th><th>Nota</th><th>{anio}</th><th>{int(anio)-1}</th></tr>"
    tot_c_y2, tot_c_y1, tot_nc_y2, tot_nc_y1 = 0, 0, 0, 0
    if tipo == 'activo':
        html += "<tr><td class='seccion' colspan='4'>Activos corrientes:</td></tr>"
        for _, r in df_comp[df_comp['codigo'].str.startswith('1', na=False)].iterrows():
            if not es_activo_no_corriente(r['codigo'], r['cuenta']):
                v2, v1 = abs(r['saldo_final_Y2']), abs(r['saldo_final_Y1'])
                if v2 == 0 and v1 == 0: continue
                tot_c_y2 += v2; tot_c_y1 += v1
                html += f"<tr><td>{r['cuenta'].title()}</td><td></td><td>{fmt_c(v2)}</td><td>{fmt_c(v1)}</td></tr>"
        html += f"<tr class='subtotal'><td>Total activos corrientes</td><td></td><td>{fmt_c(tot_c_y2)}</td><td>{fmt_c(tot_c_y1)}</td></tr>"
        html += "<tr><td class='seccion' colspan='4'>Activos no corrientes:</td></tr>"
        for _, r in df_comp[df_comp['codigo'].str.startswith('1', na=False)].iterrows():
            if es_activo_no_corriente(r['codigo'], r['cuenta']):
                v2, v1 = abs(r['saldo_final_Y2']), abs(r['saldo_final_Y1'])
                if 'acum' in str(r['cuenta']).lower(): v2, v1 = -v2, -v1 
                if v2 == 0 and v1 == 0: continue
                tot_nc_y2 += v2; tot_nc_y1 += v1
                html += f"<tr><td>{r['cuenta'].title()}</td><td></td><td>{fmt_c(v2)}</td><td>{fmt_c(v1)}</td></tr>"
        html += f"<tr class='subtotal'><td>Total activos no corrientes</td><td></td><td>{fmt_c(tot_nc_y2)}</td><td>{fmt_c(tot_nc_y1)}</td></tr>"
    else:
        html += "<tr><td class='seccion' colspan='4'>Pasivos corrientes:</td></tr>"
        for _, r in df_comp[df_comp['codigo'].str.startswith('2', na=False)].iterrows():
            if not es_pasivo_no_corriente(r['codigo'], r['cuenta']):
                v2, v1 = abs(r['saldo_final_Y2']), abs(r['saldo_final_Y1'])
                if v2 == 0 and v1 == 0: continue
                tot_c_y2 += v2; tot_c_y1 += v1
                html += f"<tr><td>{r['cuenta'].title()}</td><td></td><td>{fmt_c(v2)}</td><td>{fmt_c(v1)}</td></tr>"
        html += f"<tr class='subtotal'><td>Total pasivos corrientes</td><td></td><td>{fmt_c(tot_c_y2)}</td><td>{fmt_c(tot_c_y1)}</td></tr>"
        html += "<tr><td class='seccion' colspan='4'>Pasivos no corrientes:</td></tr>"
        for _, r in df_comp[df_comp['codigo'].str.startswith('2', na=False)].iterrows():
            if es_pasivo_no_corriente(r['codigo'], r['cuenta']):
                v2, v1 = abs(r['saldo_final_Y2']), abs(r['saldo_final_Y1'])
                if v2 == 0 and v1 == 0: continue
                tot_nc_y2 += v2; tot_nc_y1 += v1
                html += f"<tr><td>{r['cuenta'].title()}</td><td></td><td>{fmt_c(v2)}</td><td>{fmt_c(v1)}</td></tr>"
        html += f"<tr class='subtotal'><td>Total pasivos no corrientes</td><td></td><td>{fmt_c(tot_nc_y2)}</td><td>{fmt_c(tot_nc_y1)}</td></tr>"
        pat_y2, pat_y1 = 0, 0
        html += "<tr><td class='seccion' colspan='4'>Patrimonio:</td></tr>"
        for _, r in df_comp[df_comp['codigo'].str.startswith('3', na=False)].iterrows():
            v2, v1 = abs(r['saldo_final_Y2']), abs(r['saldo_final_Y1'])
            if v2 == 0 and v1 == 0: continue
            pat_y2 += v2; pat_y1 += v1
            html += f"<tr><td>{r['cuenta'].title()}</td><td></td><td>{fmt_c(v2)}</td><td>{fmt_c(v1)}</td></tr>"
        html += f"<tr class='subtotal'><td>Total Patrimonio</td><td></td><td>{fmt_c(pat_y2)}</td><td>{fmt_c(pat_y1)}</td></tr>"
    gran_tot_y2 = tot_c_y2 + tot_nc_y2 + (pat_y2 if tipo == 'pasivo' else 0)
    gran_tot_y1 = tot_c_y1 + tot_nc_y1 + (pat_y1 if tipo == 'pasivo' else 0)
    titulo_tot = "Total Activos" if tipo == 'activo' else "Total Pasivos y Patrimonio"
    html += f"<tr class='total'><td>{titulo_tot}</td><td></td><td>{fmt_c(gran_tot_y2)}</td><td>{fmt_c(gran_tot_y1)}</td></tr></table>"
    return html

def html_flujo_hoja_trabajo(df_comp, anio):
    html = f"<table class='tabla-contable'><tr><th>Hoja de Flujo de Efectivo</th><th>{anio}</th><th>{int(anio)-1}</th></tr>"
    html += "<tr><td class='seccion' colspan='3'>Activos</td></tr>"
    for _, r in df_comp[df_comp['codigo'].str.startswith('1', na=False)].iterrows():
        v2, v1 = abs(r['saldo_final_Y2']), abs(r['saldo_final_Y1'])
        if 'acum' in str(r['cuenta']).lower(): v2, v1 = -v2, -v1
        if v2 != 0 or v1 != 0: html += f"<tr><td>{r['cuenta'].title()}</td><td>{fmt_c(v2)}</td><td>{fmt_c(v1)}</td></tr>"
    html += "<tr><td class='seccion' colspan='3'>Pasivos</td></tr>"
    for _, r in df_comp[df_comp['codigo'].str.startswith('2', na=False)].iterrows():
        v2, v1 = -abs(r['saldo_final_Y2']), -abs(r['saldo_final_Y1'])
        if v2 != 0 or v1 != 0: html += f"<tr><td>{r['cuenta'].title()}</td><td>{fmt_c(v2)}</td><td>{fmt_c(v1)}</td></tr>"
    return html + "</table>"

def html_borrador_ir2(df_bal, periodo):
    ingresos         = abs(df_bal[df_bal['codigo'].str.startswith('4', na=False)]['saldo_final'].sum())
    costos           = abs(df_bal[df_bal['codigo'].str.startswith('5', na=False)]['saldo_final'].sum())
    gastos           = abs(df_bal[df_bal['codigo'].str.startswith('6', na=False)]['saldo_final'].sum())
    utilidad_neta    = ingresos - costos - gastos
    rni              = max(0, utilidad_neta)
    isr_liquidado    = rni * 0.27
    activos_totales  = abs(df_bal[df_bal['codigo'].str.startswith('1', na=False)]['saldo_final'].sum())
    impuesto_activos = activos_totales * 0.01
    impuesto_mayor   = max(isr_liquidado, impuesto_activos)

    depreciacion_libros = abs(df_bal[df_bal['cuenta'].str.lower().str.contains('deprecia', na=False)]['saldo_final'].sum())
    inventario          = abs(df_bal[df_bal['cuenta'].str.lower().str.contains('inventari', na=False)]['saldo_final'].sum())
    provisiones_nc      = abs(df_bal[df_bal['cuenta'].str.lower().str.contains('provision', na=False)]['saldo_final'].sum())

    html = "<table class='tabla-ir2'>"
    html += f"<tr><th colspan='3'>DECLARACIÓN JURADA ANUAL DEL IMPUESTO SOBRE LA RENTA DE SOCIEDADES (IR-2)<br><span style='font-weight:normal;font-size:0.85rem;'>PERÍODO FISCAL: {periodo}</span></th></tr>"
    html += "<tr><td colspan='3' class='header-seccion'>📋 ANEXO B-1 — ESTADO DE RESULTADOS (Base Fiscal)</td></tr>"
    html += f"<tr><td class='col-num'>1.1</td><td class='col-desc'>Ingresos Ventas Locales (Casilla 1.1)</td><td class='col-monto'>RD$ {ingresos:,.2f}</td></tr>"
    html += f"<tr><td class='col-num'>5</td><td class='col-desc'>Costo de Ventas (Viene del Anexo D, Casilla 39)</td><td class='col-monto' style='color:#dc2626;'>RD$ ({costos:,.2f})</td></tr>"
    html += f"<tr><td class='col-num'>6</td><td class='col-desc'>Gastos de Personal y Operacionales</td><td class='col-monto' style='color:#dc2626;'>RD$ ({gastos:,.2f})</td></tr>"
    html += f"<tr class='fila-total'><td class='col-num'>14</td><td class='col-desc'>Beneficio (Pérdida) antes del ISR</td><td class='col-monto'>RD$ {utilidad_neta:,.2f}</td></tr>"
    html += "<tr><td colspan='3' class='header-seccion'>⚙️ ANEXO G — AJUSTES FISCALES</td></tr>"
    html += f"<tr><td class='col-num'>G.1</td><td class='col-desc'>Depreciación contabilizada en libros</td><td class='col-monto'>RD$ {depreciacion_libros:,.2f}</td></tr>"
    html += f"<tr><td class='col-num'>G.2</td><td class='col-desc'>Inventarios al cierre del período</td><td class='col-monto'>RD$ {inventario:,.2f}</td></tr>"
    html += f"<tr><td class='col-num'>G.3</td><td class='col-desc'>Provisiones registradas (validar admisibilidad Art.288)</td><td class='col-monto' style='color:#f59e0b;'>RD$ {provisiones_nc:,.2f}</td></tr>"
    html += "<tr><td colspan='3' class='header-seccion'>II. DETERMINACIÓN DE LA RENTA NETA IMPONIBLE</td></tr>"
    html += f"<tr><td class='col-num'>1</td><td class='col-desc'>Beneficio o Pérdida Neta antes del ISR (Casilla 1 IR-2)</td><td class='col-monto'>RD$ {utilidad_neta:,.2f}</td></tr>"
    html += f"<tr><td class='col-num'>7</td><td class='col-desc'>Renta Neta Imponible antes de pérdidas (Casillas 1 ± 6)</td><td class='col-monto'>RD$ {rni:,.2f}</td></tr>"
    html += f"<tr class='fila-total'><td class='col-num'>11</td><td class='col-desc'>Renta Neta Imponible Final (Base ISR)</td><td class='col-monto'>RD$ {rni:,.2f}</td></tr>"
    html += "<tr><td colspan='3' class='header-seccion'>III. LIQUIDACIÓN DEL IMPUESTO SOBRE LA RENTA</td></tr>"
    html += f"<tr class='fila-total'><td class='col-num'>12</td><td class='col-desc'>Impuesto Liquidado (27% × Casilla 11)</td><td class='col-monto'>RD$ {isr_liquidado:,.2f}</td></tr>"
    html += f"<tr><td class='col-num'>13</td><td class='col-desc'>Anticipos Pagados (completar manualmente)</td><td class='col-monto' style='color:#64748b;'>RD$ 0.00</td></tr>"
    html += f"<tr><td class='col-num'>23</td><td class='col-desc'>Diferencia a Pagar (si positivo)</td><td class='col-monto'>RD$ {isr_liquidado:,.2f}</td></tr>"
    html += "<tr><td colspan='3' class='header-seccion'>IV. LIQUIDACIÓN DEL IMPUESTO A LOS ACTIVOS (Formulario ACTIVO)</td></tr>"
    html += f"<tr><td class='col-num'>34</td><td class='col-desc'>Activos Imponibles (Total Activos)</td><td class='col-monto'>RD$ {activos_totales:,.2f}</td></tr>"
    html += f"<tr><td class='col-num'>39</td><td class='col-desc'>Impuesto a los Activos (1% × Casilla 38)</td><td class='col-monto'>RD$ {impuesto_activos:,.2f}</td></tr>"
    html += f"<tr><td class='col-num'>40</td><td class='col-desc'>Crédito ISR Liquidado (viene Casilla 12)</td><td class='col-monto' style='color:#dc2626;'>RD$ ({isr_liquidado:,.2f})</td></tr>"
    dif_activos = max(0, impuesto_activos - isr_liquidado)
    html += f"<tr><td class='col-num'>41</td><td class='col-desc'>Diferencia a Pagar Impuesto Activos (si positivo)</td><td class='col-monto'>RD$ {dif_activos:,.2f}</td></tr>"
    html += "<tr><td colspan='3' class='header-seccion'>V. RESUMEN — MAYOR OBLIGACIÓN TRIBUTARIA</td></tr>"
    html += f"<tr class='fila-total' style='background-color:#1e3a5f; color:white;'><td class='col-num' style='background-color:#1e3a5f; color:white;'>★</td><td class='col-desc'>IMPUESTO MAYOR A PAGAR (Art. 314 CTRD: mayor entre ISR e Impuesto Activos)</td><td class='col-monto'>RD$ {impuesto_mayor:,.2f}</td></tr>"
    html += "</table>"
    return html

def html_it1_detallado(monto_607, itbis_607, monto_606, itbis_606, df_607, df_606, periodo):
    itbis_neto = itbis_607 - itbis_606
    proporcionalidad_pct = (itbis_607 / monto_607 * 100) if monto_607 > 0 else 0

    html = "<table class='tabla-ir2'>"
    html += f"<tr><th colspan='3'>PRELIQUIDACIÓN IT-1 — IMPUESTO A LAS TRANSFERENCIAS DE BIENES INDUSTRIALIZADOS Y SERVICIOS<br><span style='font-weight:normal;font-size:0.85rem;'>PERÍODO: {periodo}</span></th></tr>"
    html += "<tr><td colspan='3' class='header-seccion'>📊 SECCIÓN I — OPERACIONES DE VENTAS (Formato 607)</td></tr>"
    html += f"<tr><td class='col-num'>1.1</td><td class='col-desc'>Total Monto Facturado (Ventas)</td><td class='col-monto'>RD$ {monto_607:,.2f}</td></tr>"
    html += f"<tr><td class='col-num'>1.2</td><td class='col-desc'>ITBIS Facturado / Cobrado (18%)</td><td class='col-monto'>RD$ {itbis_607:,.2f}</td></tr>"
    html += f"<tr><td class='col-num'>1.3</td><td class='col-desc'>Tasa Efectiva ITBIS sobre ventas</td><td class='col-monto'>{proporcionalidad_pct:.2f}%</td></tr>"
    html += "<tr><td colspan='3' class='header-seccion'>🛒 SECCIÓN II — OPERACIONES DE COMPRAS (Formato 606)</td></tr>"
    html += f"<tr><td class='col-num'>2.1</td><td class='col-desc'>Total Monto Compras y Servicios</td><td class='col-monto'>RD$ {monto_606:,.2f}</td></tr>"
    html += f"<tr><td class='col-num'>2.2</td><td class='col-desc'>ITBIS Adelantado (por adelantar)</td><td class='col-monto'>RD$ {itbis_606:,.2f}</td></tr>"
    html += "<tr><td colspan='3' class='header-seccion'>⚖️ SECCIÓN III — DETERMINACIÓN ITBIS A PAGAR</td></tr>"
    html += f"<tr><td class='col-num'>3.1</td><td class='col-desc'>ITBIS Cobrado en Ventas</td><td class='col-monto'>RD$ {itbis_607:,.2f}</td></tr>"
    html += f"<tr><td class='col-num'>3.2</td><td class='col-desc'>(-) ITBIS Adelantado en Compras</td><td class='col-monto' style='color:#dc2626;'>RD$ ({itbis_606:,.2f})</td></tr>"
    label = "ITBIS A PAGAR" if itbis_neto > 0 else "SALDO A FAVOR"
    html += f"<tr class='fila-total' style='background-color:#1e3a5f; color:white;'><td class='col-num' style='background-color:#1e3a5f;color:white;'>★</td><td class='col-desc'>{label}</td><td class='col-monto'>RD$ {abs(itbis_neto):,.2f}</td></tr>"
    html += "</table>"
    
    if monto_607 > 0 and proporcionalidad_pct > 0 and abs(proporcionalidad_pct - 18) > 3:
        html += f"<div class='alerta-box'>⚠️ La tasa efectiva de ITBIS sobre ventas ({proporcionalidad_pct:.1f}%) difiere significativamente del 18%. Verificar operaciones exentas o gravadas a tasa reducida declaradas en el 607.</div>"
    if itbis_neto < 0:
        html += f"<div class='ok-box'>✅ Saldo a favor de RD$ {abs(itbis_neto):,.2f}. Puede solicitar compensación o devolución ante la DGII conforme al Art. 50 del Código Tributario.</div>"
    elif itbis_neto > 0:
        html += f"<div class='info-box'>ℹ️ ITBIS a pagar de RD$ {itbis_neto:,.2f}. Verificar fecha límite de presentación del IT-1 (día 20 del mes siguiente).</div>"
    return html

# ──────────────────────────────────────────────────────────────────────────────
# SIDEBAR
# ──────────────────────────────────────────────────────────────────────────────
st.sidebar.title("🏛️ TaxTech Auditor RD")
st.sidebar.markdown("---")
empresa = st.sidebar.text_input("Empresa", value="Empresa de Prueba SRL")
periodo = st.sidebar.text_input("Período", value="Enero - Diciembre 2026")
anio    = st.sidebar.text_input("Año Fiscal", value="2026")

st.sidebar.markdown("---")
st.sidebar.markdown("### 📂 Estados Financieros")
uploaded      = st.sidebar.file_uploader("Balanza (Año Actual)",   type=["xlsx","xls","csv"])
uploaded_prev = st.sidebar.file_uploader("Balanza (Año Anterior)", type=["xlsx","xls","csv"])

st.sidebar.markdown("---")
st.sidebar.markdown("### 🧾 Archivos Fiscales (independientes)")
file_606 = st.sidebar.file_uploader("Formato 606 (Compras DGII)", type=["xlsx","xls","csv","txt"])
file_607 = st.sidebar.file_uploader("Formato 607 (Ventas DGII)",  type=["xlsx","xls","csv","txt"])
file_tss = st.sidebar.file_uploader("Plantilla TSS",              type=["xlsx","xls","csv"])

st.sidebar.markdown("---")
st.sidebar.markdown("### ⚙️ Parámetros TSS")
techo_sfs_input = st.sidebar.number_input("Techo SFS mensual (RD$)", value=TECHO_SFS, step=1000.0, format="%.2f")
techo_afp_input = st.sidebar.number_input("Techo AFP mensual (RD$)", value=TECHO_AFP, step=1000.0, format="%.2f")
TECHO_SFS = techo_sfs_input; TECHO_AFP = techo_afp_input

# ──────────────────────────────────────────────────────────────────────────────
# MAIN
# ──────────────────────────────────────────────────────────────────────────────
st.title("TaxTech Auditor — Declaraciones Juradas & Estados Financieros")

if not uploaded and not file_606 and not file_607 and not file_tss:
    st.markdown("""
    <div class='info-box'>
    👆 <strong>Carga al menos un archivo</strong> para iniciar tu auditoría:<br>
    • <strong>Balanza de comprobación</strong> → genera todos los EEFF y el borrador IR-2<br>
    • <strong>Formato 606/607</strong> → análisis IT-1 de ITBIS de forma independiente<br>
    • <strong>Plantilla TSS</strong> → auditoría de nómina y cálculo de aportes independiente
    </div>
    """, unsafe_allow_html=True)
    st.stop()

df_bal  = pd.DataFrame()
df_comp = pd.DataFrame()

if uploaded:
    with st.spinner("Procesando balanza contable..."):
        df_bal = procesar_balanza(uploaded)
        if not df_bal.empty:
            df_bal = analizar_balanza(df_bal)
            if uploaded_prev:
                df_prev = procesar_balanza(uploaded_prev)
                df_comp = procesar_comparativo(df_bal, df_prev) if not df_prev.empty else pd.DataFrame()
            if df_comp.empty:
                df_comp = df_bal.copy()
                df_comp.rename(columns={'saldo_final': 'saldo_final_Y2'}, inplace=True)
                df_comp['saldo_final_Y1'] = 0.0
                df_comp['variacion_abs']  = df_comp['saldo_final_Y2']

t_ingresos = abs(df_bal[df_bal['codigo'].str.startswith('4', na=False)]['saldo_final'].sum()) if not df_bal.empty else 0
t_activos  = abs(df_bal[df_bal['codigo'].str.startswith('1', na=False)]['saldo_final'].sum()) if not df_bal.empty else 0
t_costos   = abs(df_bal[df_bal['codigo'].str.startswith('5', na=False)]['saldo_final'].sum()) if not df_bal.empty else 0
t_gastos   = abs(df_bal[df_bal['codigo'].str.startswith('6', na=False)]['saldo_final'].sum()) if not df_bal.empty else 0
utilidad_neta = t_ingresos - t_costos - t_gastos

monto_606, itbis_606, df_606_det = procesar_606_607(file_606, "606") if file_606 else (0, 0, pd.DataFrame())
monto_607, itbis_607, df_607_det = procesar_606_607(file_607, "607") if file_607 else (0, 0, pd.DataFrame())
df_tss, tss_res = procesar_tss(file_tss) if file_tss else (None, None)

st.markdown(f"### 📌 {empresa} — {periodo}")

tab_comp, tab_bg, tab_er, tab_pat, tab_efe, tab_bal, tab_inconsist, tab_art287, tab_ir2, tab_it1, tab_tss_tab, tab_consol = st.tabs([
    "📈 Dashboard", "📊 Balance General", "📉 Estado de Resultados",
    "💼 Cambios Patrimonio", "🌊 Flujo de Efectivo",
    "📋 Balanza", "🚨 Inconsistencias", "⚖️ Riesgos Art.287",
    "📝 Borrador IR-2", "🧾 IT-1 ITBIS", "👥 TSS Nómina", "🏛️ Consolidado Fiscal"
])

try:
    with tab_comp:
        if not df_comp.empty:
            c1, c2, c3, c4 = st.columns(4)
            c1.metric("Ingresos", f"RD$ {t_ingresos:,.0f}")
            c2.metric("Utilidad Neta", f"RD$ {utilidad_neta:,.0f}")
            c3.metric("Total Activos", f"RD$ {t_activos:,.0f}")
            mg = (utilidad_neta / t_ingresos * 100) if t_ingresos > 0 else 0
            c4.metric("Margen Neto", f"{mg:.1f}%")
            st.markdown("---")
            c5, c6 = st.columns(2)
            with c5:
                excel_bytes = exportar_reporte_corporativo(empresa, periodo, anio, df_comp)
                if excel_bytes:
                    st.download_button("📥 Reporte Financiero Completo (Excel)", data=excel_bytes, file_name=f"EEFF_{empresa.replace(' ','_')}_{anio}.xlsx")
            df_chart = pd.DataFrame({
                'Año':     [str(int(anio)-1), anio],
                'Ingresos': [abs(df_comp[df_comp['codigo'].str.startswith('4',na=False)]['saldo_final_Y1'].sum()), t_ingresos],
                'Activos':  [abs(df_comp[df_comp['codigo'].str.startswith('1',na=False)]['saldo_final_Y1'].sum()), t_activos],
            })
            st.plotly_chart(px.bar(df_chart, x='Año', y=['Ingresos','Activos'], barmode='group', title="Comparativo"), use_container_width=True)
        else:
            st.info("Carga la balanza de comprobación para ver el Dashboard.")

    with tab_bg:
        if not df_comp.empty:
            c1, c2 = st.columns([3, 1])
            with c1:
                col_a, col_p = st.columns(2)
                with col_a: st.markdown(html_balance_general(df_comp, anio, 'activo'), unsafe_allow_html=True)
                with col_p: st.markdown(html_balance_general(df_comp, anio, 'pasivo'), unsafe_allow_html=True)
            with c2:
                bg_bytes = exportar_balance_excel(df_comp, empresa, anio)
                st.download_button("📥 Descargar Balance (Excel)", data=bg_bytes, file_name=f"Balance_{empresa.replace(' ','_')}_{anio}.xlsx")
        else:
            st.info("Carga la balanza para generar el Balance General.")

    with tab_er:
        if not df_comp.empty:
            c1, c2 = st.columns([3, 1])
            with c1: st.markdown(html_estado_resultados(df_comp, anio), unsafe_allow_html=True)
            with c2:
                er_bytes = exportar_er_excel(df_comp, empresa, anio)
                st.download_button("📥 Descargar E.R. (Excel)", data=er_bytes, file_name=f"EstadoResultados_{empresa.replace(' ','_')}_{anio}.xlsx")
        else:
            st.info("Carga la balanza para generar el Estado de Resultados.")

    with tab_pat:
        if not df_comp.empty:
            st.markdown(html_cambios_patrimonio(df_comp, anio), unsafe_allow_html=True)
            df_pat = df_comp[df_comp['codigo'].str.startswith('3', na=False)][['codigo','cuenta','saldo_final_Y2','saldo_final_Y1']]
            st.download_button("📥 Descargar Patrimonio (Excel)", data=generar_excel_descargable(df_pat, "Patrimonio"), file_name=f"Patrimonio_{anio}.xlsx")
        else:
            st.info("Carga la balanza para generar el Estado de Cambios en el Patrimonio.")

    with tab_efe:
        if not df_comp.empty:
            st.markdown(html_flujo_hoja_trabajo(df_comp, anio), unsafe_allow_html=True)
            st.info("ℹ️ El Flujo de Efectivo completo (método indirecto) está incluido en el Reporte Financiero Completo (Excel).")
        else:
            st.info("Carga la balanza para generar el Flujo de Efectivo.")

    with tab_bal:
        if not df_bal.empty:
            st.dataframe(df_bal[['codigo','cuenta','saldo_final']], use_container_width=True)
            st.download_button("📥 Descargar Balanza (Excel)", data=generar_excel_descargable(df_bal[['codigo','cuenta','saldo_final']], "Balanza"), file_name="Balanza_Auditoria.xlsx")
        else:
            st.info("Carga la balanza de comprobación.")

    with tab_inconsist:
        if not df_bal.empty:
            df_show_inc = df_bal[~df_bal['validacion_naturaleza'].str.startswith('✅')][['codigo','cuenta','saldo_final','validacion_naturaleza']]
            if not df_show_inc.empty:
                st.warning(f"Se encontraron {len(df_show_inc)} cuentas con saldo de naturaleza contraria.")
                st.dataframe(df_show_inc, use_container_width=True)
                st.download_button("📥 Descargar (Excel)", data=generar_excel_descargable(df_show_inc), file_name="Inconsistencias.xlsx", key="btn_inc")
            else:
                st.success("✅ Sin inconsistencias de naturaleza contable detectadas.")
        else:
            st.info("Carga la balanza para buscar inconsistencias.")

    with tab_art287:
        if not df_bal.empty:
            df_show_art = df_bal[df_bal['alerta_fiscal'] != ""][['codigo','cuenta','saldo_final','alerta_fiscal']]
            if not df_show_art.empty:
                st.warning(f"Se encontraron {len(df_show_art)} cuentas con alertas de deducibilidad fiscal.")
                st.dataframe(df_show_art, use_container_width=True)
                st.download_button("📥 Descargar (Excel)", data=generar_excel_descargable(df_show_art), file_name="Riesgos_Art287.xlsx", key="btn_art")
            else:
                st.success("✅ Sin alertas de riesgo de deducibilidad detectadas.")
        else:
            st.info("Carga la balanza para validar riesgos del Art. 287.")

    with tab_ir2:
        if not df_bal.empty:
            st.markdown(html_borrador_ir2(df_bal, periodo), unsafe_allow_html=True)
            st.markdown("---")
            isr_est = max(0, utilidad_neta) * 0.27
            df_ir2_export = pd.DataFrame([
                ("PERÍODO FISCAL", periodo),
                ("Total Ingresos (Casilla A)", f"RD$ {t_ingresos:,.2f}"),
                ("Costos de Ventas", f"RD$ {t_costos:,.2f}"),
                ("Gastos Operacionales", f"RD$ {t_gastos:,.2f}"),
                ("Beneficio/Pérdida antes ISR (Cas.1)", f"RD$ {utilidad_neta:,.2f}"),
                ("Renta Neta Imponible (Cas.11)", f"RD$ {max(0,utilidad_neta):,.2f}"),
                ("ISR Liquidado 27% (Cas.12)", f"RD$ {isr_est:,.2f}"),
                ("Total Activos Imponibles (Cas.34)", f"RD$ {t_activos:,.2f}"),
                ("Impuesto Activos 1% (Cas.39)", f"RD$ {t_activos*0.01:,.2f}"),
                ("IMPUESTO MAYOR A PAGAR", f"RD$ {max(isr_est, t_activos*0.01):,.2f}"),
            ], columns=["Concepto","Valor"])
            st.download_button("📥 Descargar Borrador IR-2 (Excel)", data=generar_excel_descargable(df_ir2_export, "Borrador IR-2"), file_name=f"BorradorIR2_{empresa.replace(' ','_')}_{anio}.xlsx")
        else:
            st.info("Carga la balanza para generar el borrador del IR-2.")

    with tab_it1:
        st.markdown("### 🧾 Preliquidación IT-1 — ITBIS")
        if not file_606 and not file_607:
            st.markdown("""
            <div class='info-box'>
            👆 Este módulo funciona de forma <strong>completamente independiente</strong> de la balanza.<br>
            Carga el <strong>Formato 606</strong> (compras) y/o <strong>607</strong> (ventas) exportados desde la OFV de la DGII.
            </div>
            """, unsafe_allow_html=True)
        else:
            c1, c2, c3, c4 = st.columns(4)
            c1.metric("Ventas (607)", f"RD$ {monto_607:,.0f}")
            c2.metric("ITBIS Cobrado", f"RD$ {itbis_607:,.0f}")
            c3.metric("Compras (606)", f"RD$ {monto_606:,.0f}")
            c4.metric("ITBIS Adelantado", f"RD$ {itbis_606:,.0f}")
            st.markdown("---")
            st.markdown(html_it1_detallado(monto_607, itbis_607, monto_606, itbis_606, df_607_det, df_606_det, periodo), unsafe_allow_html=True)
            st.markdown("---")
            if not df_607_det.empty:
                with st.expander("📋 Ver registros del Formato 607 (Ventas)"):
                    st.dataframe(df_607_det, use_container_width=True)
            if not df_606_det.empty:
                with st.expander("🛒 Ver registros del Formato 606 (Compras)"):
                    st.dataframe(df_606_det, use_container_width=True)
            it1_bytes = exportar_it1_excel(empresa, periodo, monto_607, itbis_607, monto_606, itbis_606, df_607_det, df_606_det)
            st.download_button("📥 Descargar Preliquidación IT-1 (Excel)", data=it1_bytes, file_name=f"PreliquidacionIT1_{empresa.replace(' ','_')}_{periodo.replace(' ','_')}.xlsx")

    with tab_tss_tab:
        st.markdown("### 👥 Auditoría de Nómina y TSS")
        st.download_button("📥 Descargar Plantilla Vacía Autodeterminación", data=generar_plantilla_tss(), file_name="Plantilla_TSS_Vacia.xlsx")
        st.markdown("""
        <div class='info-box'>
        Este módulo funciona de forma <strong>completamente independiente</strong> de la balanza.<br>
        Carga la plantilla de Autodeterminación TSS para calcular los aportes de cada empleado con los techos vigentes.
        </div>
        """, unsafe_allow_html=True)
        if file_tss and tss_res:
            c1, c2, c3, c4 = st.columns(4)
            c1.metric("Empleados Validados", tss_res['total_empleados'])
            c2.metric("Nómina Mensual", f"RD$ {tss_res['nomina_mensual']:,.2f}")
            c3.metric("Aportes Patronales", f"RD$ {tss_res['total_patronal']:,.2f}")
            c4.metric("Retenciones Empleados", f"RD$ {tss_res['total_empleado']:,.2f}")
            st.markdown("---")
            with st.expander("📊 Ver desglose de tasas aplicadas"):
                st.markdown(f"""
                | Concepto | Tasa | Base |
                |---|---|---|
                | SFS Patronal | {TASA_SFS_PAT*100:.2f}% | Techo RD$ {TECHO_SFS:,.2f} |
                | AFP Patronal | {TASA_AFP_PAT*100:.2f}% | Techo RD$ {TECHO_AFP:,.2f} |
                | SRL Patronal | {TASA_SRL*100:.2f}% | Salario bruto |
                | INFOTEP | {TASA_INFOTEP*100:.2f}% | Salario bruto |
                | SFS Empleado | {TASA_SFS_EMP*100:.2f}% | Techo RD$ {TECHO_SFS:,.2f} |
                | AFP Empleado | {TASA_AFP_EMP*100:.2f}% | Techo RD$ {TECHO_AFP:,.2f} |
                """)
            st.markdown(f"**Total mensual a remitir TSS:** `RD$ {tss_res['total_pagar']:,.2f}`")
            if not df_bal.empty:
                gastos_personal_bal = abs(df_bal[df_bal['cuenta'].str.lower().str.contains('tss|seguridad social|aporte pat', na=False)]['saldo_final'].sum())
                if gastos_personal_bal > 0:
                    dif_cruce = abs(tss_res['total_patronal'] - gastos_personal_bal)
                    if dif_cruce > 100:
                        st.markdown(f"<div class='alerta-box'>⚠️ Diferencia en cruce TSS vs contabilidad: RD$ {dif_cruce:,.2f}. Verificar registros de aportes patronales en la balanza.</div>", unsafe_allow_html=True)
                    else:
                        st.markdown("<div class='ok-box'>✅ Aportes TSS conciliados con la contabilidad.</div>", unsafe_allow_html=True)
            st.dataframe(df_tss, use_container_width=True)
            tss_bytes = exportar_tss_excel(df_tss, tss_res, empresa, periodo)
            st.download_button("📥 Descargar Auditoría TSS (Excel)", data=tss_bytes, file_name=f"AuditoriaTSS_{empresa.replace(' ','_')}_{periodo.replace(' ','_')}.xlsx")
        elif file_tss:
            st.error("No se pudo procesar el archivo TSS. Verifica que tenga la columna 'Salario Ordinario' o 'Sueldo'.")

    with tab_consol:
        st.markdown("### 🏛️ Resumen Fiscal Consolidado")
        isr_est = max(0, utilidad_neta) * 0.27

        estado_ir2 = "✅ Base Balanza" if not df_bal.empty else "⚠️ Falta Balanza"
        estado_it1 = "✅ Calculado (606/607)" if (file_606 or file_607) else "⚠️ Falta 606/607"
        estado_tss = "✅ Calculado (Plantilla)" if file_tss else "⚠️ Falta Plantilla TSS"

        c1, c2, c3 = st.columns(3)
        c1.metric("ISR Estimado (IR-2)", f"RD$ {isr_est:,.0f}", help="27% sobre Renta Neta Imponible")
        c2.metric("ITBIS Neto (IT-1)", f"RD$ {(itbis_607-itbis_606):,.0f}", delta=f"{'Pagar' if itbis_607>itbis_606 else 'A favor'}")
        c3.metric("TSS Mensual", f"RD$ {tss_res['total_pagar'] if tss_res else 0:,.0f}")

        st.markdown("---")
        df_consol = pd.DataFrame([
            ("IT-1 / ITBIS",  "Impuesto Transferencias (mensual)",    f"RD$ {(itbis_607-itbis_606):,.2f}", estado_it1),
            ("TSS",           "Aportes y Retenciones Nómina",          f"RD$ {tss_res['total_pagar'] if tss_res else 0:,.2f}", estado_tss),
            ("IR-2 / ISR",    "Impuesto Sobre la Renta (anual)",       f"RD$ {isr_est:,.2f}", estado_ir2),
            ("ACTIVO",        "Impuesto a los Activos (anual)",        f"RD$ {t_activos*0.01:,.2f}", estado_ir2),
            ("MAYOR",         "Mayor Obligación IR-2 / ACTIVO",        f"RD$ {max(isr_est, t_activos*0.01):,.2f}", estado_ir2),
        ], columns=["Formulario", "Concepto", "Monto Estimado", "Estado"])

        st.dataframe(df_consol, use_container_width=True, hide_index=True)
        st.download_button("📥 Descargar Consolidado (Excel)", data=generar_excel_descargable(df_consol, "Consolidado Fiscal"), file_name=f"ConsolidadoFiscal_{empresa.replace(' ','_')}_{anio}.xlsx")

        if not df_bal.empty:
            st.markdown("---")
            st.markdown("#### Indicadores Financieros Clave")
            c1, c2, c3, c4 = st.columns(4)
            t_pasivos = abs(df_bal[df_bal['codigo'].str.startswith('2', na=False)]['saldo_final'].sum())
            t_pat = t_activos - t_pasivos
            c1.metric("Liquidez Corriente", f"{(t_activos/t_pasivos if t_pasivos else 0):.2f}x")
            c2.metric("Deuda / Patrimonio",  f"{(t_pasivos/t_pat if t_pat else 0):.2f}x")
            c3.metric("Margen Bruto", f"{((t_ingresos-t_costos)/t_ingresos*100 if t_ingresos else 0):.1f}%")
            c4.metric("ROA", f"{(utilidad_neta/t_activos*100 if t_activos else 0):.1f}%")

except Exception as e:
    st.error(f"Error al renderizar la interfaz: {e}")